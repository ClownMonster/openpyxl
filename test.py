
import openpyxl
import os
import sys

wb = " " 
class pyxl:
  
  #constructor of the class pyxl
  def __init__(self, path):
    self.path = path
    print(os.path.abspath(path))
    pass

  #method to load the sheet 
  def load_sheet(self):`
    global wb
    wb = openpyxl.load_workbook(self.path) # loding your xlsx sheet to your wb
    sheet = wb.active # activating wb to start accessing  data from it
    return sheet

  def save_the_file(self):
    global wb
    wb.save("mohan.xlsx") # Saving the changes you have made to your file
  pass
#End of the class



def data_entry(S):
  row_no = int(input("Enter the Row number where you need to enter the data : ")) # hold the row to place the data
  col_no = int(input("Enter the column number where you need to enter the data : ")) # hold the column to place the data
  if row_no <=0  or  col_no <= 0:
    print("The row or column value Doesn't exits")
    data_entry(S)
  S.cell(row = row_no, column = col_no ).value = str(input("Enter the value tobe added :")) # assigning the value of the respective cell
  return


def print_data(S):
  row_no = int(input("Enter the Value of the Row to find the data : "))
  col_no = int(input("Enter the Value of the Row to find the data : "))
  if row_no <=0 or col_no <= 0:
    print("The row value Doesn't exits")
    print_data(S)
  print(f"value : {S.cell(row = row_no , column = col_no).value}")
  return
    

#main function
if __name__== '__main__':  
  o = pyxl(input("Enter the path for the xlsx file : ")) # Taking the input path if the file is in the same folder you can just give the name of the file
  sheet =  o.load_sheet()
  while(True):
    try:
      choice = int(input("\n1.Data Entry\n2.Print Data\n3.Save and Exit\nEnter your choice : "))
    except:
      print("Invalid Value")

    if(choice == 1):
       data_entry(sheet)

    elif(choice == 2):
       print_data(sheet)

    elif(choice == 3):
      o.save_the_file()
      sys.exit(0)
    else:
      print("Invalid Choice")
pass









